from django.contrib import admin
from django.http import HttpResponse
from .models import *
from openpyxl import load_workbook,Workbook
from django.urls import path

# Register your models here.
from .models import SearchTermsFile,Dictionary,ProcessingQ

## New
from .forms import XlsxImportForm, XlsxImportSearchTermForm
from django.shortcuts import render, redirect
from openpyxl import load_workbook, Workbook, drawing
from .models import SearchDictionary, SearchPortfolio, SearchTerm, SearchClassification, SearchMatchTotal, SearchMatch
from main.classify import classify_search_terms2

def load_dict(d="/home/jakamkon/webapps/upwork_search_classification_static/media/dictonaries/Dictonary.xlsx"):
  wb=load_workbook(d)
  ws=wb.active
  st_dict={}
  for r in ws.rows:
    k=r[0].value.lower()
    k=' '.join(k.split())
    st_dict[k]=r[1].value
  return st_dict 

def classify(v, d):
  classes=[]
  v=' '.join(v.lower().split())
  keys=[]
  for k in d:
    ks=k.split()
    # For single words we want to match
    # them using spaces to make sure that
    # we match the whole word.
    if len(ks)==1:
      kl=[' %s ' % k,'%s ' % k,' %s' % k]
      for kk in kl:
        if kk in v:
          keys.append(k)
          classes.append(d[k])
    else:
      if k in v:
        keys.append(k)
        classes.append(d[k])
  classes=set(classes)
  keys=set(keys)
  return ' + '.join(list(classes)),','.join(list(keys))

def classify_search_terms(modeladmin, request, queryset):
  f=queryset[0].search_terms_file.path
  #f=open(f, 'rb')
  wb=load_workbook(f)
  ws=wb['data']
  c=ws['B']
  out_wb=Workbook()
  out_ws=out_wb.active
  #raise Exception(out_ws.sheet_state)
  cls_dict=load_dict(queryset[0].dictionary.dictionary_file.path)
  for row in c:
    st=row.value
    clsi,mkeys=classify(st, cls_dict)
    out_ws.append([st, clsi, mkeys])
  #raise Exception([ r.value for r in c])
  outf=f.split('.')[0]+'-classification.xlsx'
  out_wb.save(outf)
  return HttpResponse(open(outf,'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


classify_search_terms.short_description = "Classify search terms"

def process(modeladmin, request, queryset):
  for stf in queryset:
    pq=ProcessingQ.objects.filter(search_terms_file=stf)
    if len(pq) == 0:
      pq=ProcessingQ(search_terms_file=stf, status_info='Just added to the ProcessingQ, waiting for processing...')
      pq.save()
    else:
      pq=pq[0]
      pq.status_info='Just added to the ProcessingQ, waiting for processing...'
      pq.status='new'
      pq.save()
          
process.short_description='Process file (add it to the ProcessingQ)'

class ClassifySearchTermsAdmin(admin.ModelAdmin):
  #actions = [process]
  pass

class PQAdmin(admin.ModelAdmin):
  list_fields = ('search_terms_file_paid','status_info')

admin.site.register(SearchTermsFile,ClassifySearchTermsAdmin)
admin.site.register(ProcessingQ,PQAdmin)
admin.site.register(Dictionary)

## New
@admin.register(SearchDictionary)
class SearchDictionaryAdmin(admin.ModelAdmin):
  list_display = ('name', 'attribute')
  list_filter = ('attribute',)
  change_list_template = "admin/search_dictionary_changelist.html"

  def get_urls(self):
    urls = super().get_urls()
    my_urls = [
      path('import-xlsx/', self.import_xlsx),
    ]
    return my_urls + urls

  def import_xlsx(self, request):
    if request.method == "POST":
      xlsx_file = request.FILES["xlsx_file"]
      wb=load_workbook(xlsx_file)
      ws=wb.active
      dictionary_list = []
      for row in tuple(ws.rows)[1:]:
        dictionary_list.append(SearchDictionary(name=row[0].value, attribute=row[1].value))
      
      SearchDictionary.objects.all().delete()
      SearchDictionary.objects.bulk_create(dictionary_list)      

      self.message_user(request, "Your xlsx file has been imported")
      return redirect("..")
    form = XlsxImportForm()
    payload = {"form": form}
    return render(
      request, "admin/xlsx_form.html", payload
    )

@admin.register(SearchPortfolio)
class SearchPortfolioAdmin(admin.ModelAdmin):
  list_display = ('name',)

@admin.register(SearchTerm)
class SearchTermAdmin(admin.ModelAdmin):
  list_display = ('query', 'clicks', 'impressions', 'avg_position', 'cost', 'conversions', 'total_conv', 'search_portfolio',)
  list_filter = ('search_portfolio',)
  change_list_template = "admin/search_term_changelist.html"

  def get_urls(self):
    urls = super().get_urls()
    my_urls = [
      path('import-xlsx/', self.import_xlsx),
    ]
    return my_urls + urls

  def import_xlsx(self, request):
    if request.method == "POST":
      xlsx_file = request.FILES["xlsx_file"]
      wb=load_workbook(xlsx_file)
      ws=wb.active
      portfolio = SearchPortfolio.objects.get(pk=request.POST['portfolio'])

      if request.POST['portfolio'] == '1': # organic
        self.import_organic(ws, portfolio)
      elif request.POST['portfolio'] == '2': # paid
        self.import_paid(ws, portfolio)
      elif request.POST['portfolio'] == '3': # none
        self.import_none(ws, portfolio)

      self.message_user(request, "Your xlsx file has been imported")
      return redirect("..")
    form = XlsxImportSearchTermForm()
    payload = {"form": form}
    return render(
      request, "admin/xlsx_form.html", payload
    )

  def import_organic(self, ws, portfolio):
    terms_list = []
    for row in tuple(ws.rows)[1:]:
      terms_list.append(SearchTerm(query=row[0].value, clicks=row[1].value, impressions=row[2].value, avg_position=row[3].value, search_portfolio=portfolio))
    
    SearchTerm.objects.bulk_create(terms_list)

  def import_paid(self, ws, portfolio):
    terms_list = []
    for row in tuple(ws.rows)[1:]:
      terms_list.append(SearchTerm(query=row[0].value, impressions=row[1].value, clicks=row[2].value, cost=row[3].value, conversions=row[4].value, total_conv=row[5].value, search_portfolio=portfolio))
    
    SearchTerm.objects.bulk_create(terms_list)

  def import_none(self, ws, portfolio):
    terms_list = []
    for row in tuple(ws.rows)[1:]:
      terms_list.append(SearchTerm(query=row[0].value, search_portfolio=portfolio))
    
    SearchTerm.objects.bulk_create(terms_list)

@admin.register(SearchClassification)
class SearchClassificationAdmin(admin.ModelAdmin):
  list_display = ('query', 'search_dictionary_ids', 'portfolio_ids', 'paid_impressions', 'paid_clicks', 'paid_cost', 'paid_conversions', 'paid_total_conv', 'organic_clicks', 'organic_impressions', 'organic_avg_position',)
  change_list_template = "admin/search_classification_changelist.html"

  def get_urls(self):
    urls = super().get_urls()
    my_urls = [
      path('classify/', self.classify),
    ]
    return my_urls + urls

  def classify(self, request):
    classify_search_terms2()
    self.message_user(request, "The term classification is processing now.")
    return redirect("..")

  def changelist_view(self, request, extra_context=None):
    full = SearchMatchTotal.objects.first().full if SearchMatchTotal.objects.first() else ''
    partial = SearchMatchTotal.objects.first().partial if SearchMatchTotal.objects.first() else ''
    unclassified = SearchMatchTotal.objects.first().unclassified if SearchMatchTotal.objects.first() else ''
    unclassified_partial = SearchMatchTotal.objects.first().unclassified_partial if SearchMatchTotal.objects.first() else ''
    partial_full = SearchMatchTotal.objects.first().partial_full if SearchMatchTotal.objects.first() else ''
    unclassified_full = SearchMatchTotal.objects.first().unclassified_full if SearchMatchTotal.objects.first() else ''

    organic_full = SearchPortfolio.objects.get(name = 'organic').searchmatch_set.first().full if SearchPortfolio.objects.get(name = 'organic').searchmatch_set.first() else ''
    organic_partial = SearchPortfolio.objects.get(name = 'organic').searchmatch_set.first().partial if SearchPortfolio.objects.get(name = 'organic').searchmatch_set.first() else ''
    organic_unclassified = SearchPortfolio.objects.get(name = 'organic').searchmatch_set.first().unclassified if SearchPortfolio.objects.get(name = 'organic').searchmatch_set.first() else ''

    paid_full = SearchPortfolio.objects.get(name = 'paid').searchmatch_set.first().full if SearchPortfolio.objects.get(name = 'paid').searchmatch_set.first() else ''
    paid_partial = SearchPortfolio.objects.get(name = 'paid').searchmatch_set.first().partial if SearchPortfolio.objects.get(name = 'paid').searchmatch_set.first() else ''
    paid_unclassified = SearchPortfolio.objects.get(name = 'paid').searchmatch_set.first().unclassified if SearchPortfolio.objects.get(name = 'paid').searchmatch_set.first() else ''
    
    none_full = SearchPortfolio.objects.get(name = 'none').searchmatch_set.first().full if SearchPortfolio.objects.get(name = 'none').searchmatch_set.first() else ''
    none_partial = SearchPortfolio.objects.get(name = 'none').searchmatch_set.first().partial if SearchPortfolio.objects.get(name = 'none').searchmatch_set.first() else ''
    none_unclassified = SearchPortfolio.objects.get(name = 'none').searchmatch_set.first().unclassified if SearchPortfolio.objects.get(name = 'none').searchmatch_set.first() else ''

    my_context = {
      'full': full,
      'partial': partial,
      'unclassified': unclassified,
      'unclassified_partial': unclassified_partial,
      'partial_full': partial_full,
      'unclassified_full': unclassified_full,

      'organic_full': organic_full,
      'organic_partial': organic_partial,
      'organic_unclassified': organic_unclassified,

      'paid_full': paid_full,
      'paid_partial': paid_partial,
      'paid_unclassified': paid_unclassified,

      'none_full': none_full,
      'none_partial': none_partial,
      'none_unclassified': none_unclassified
    }
    return super(SearchClassificationAdmin, self).changelist_view(request,
      extra_context=my_context)
