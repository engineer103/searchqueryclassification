from openpyxl import load_workbook,Workbook,drawing

# Register your models here.
from .models import SearchTermsFile,Dictionary 
# from spellchecker import SpellChecker
from textblob import Word
from datetime import datetime
from .classify_v2 import classify_kws

# For venn diagram
# from matplotlib import pyplot as plt
# import PIL
# from matplotlib_venn import venn2

def load_dict(d="upwork_search_classification_static/media/dictonaries/royal_swimming_pools_dictionary.xlsx"):
   wb=load_workbook(d)
   ws=wb.active
   st_dict={}
   for r in ws.rows:
       k=r[0].value.lower()
       k=' '.join(k.split())
       st_dict[k]=r[1].value
   return st_dict 

def classify(v, d,sc=None):
   classes=[]
   kwords=str(v).lower().split()
   if sc:
       skwords=kwords[:]
       for i, kw in enumerate(kwords):
           newkw=Word(kw)
           confidence=0.0
           corrected=None
           for sc in newkw.spellcheck():
              if sc[1] < 0.9:
                 continue
              if sc[1] > confidence:
                 confidence=sc[1]
                 corrected=sc[0]
           if confidence > 0.9:
               kwords[i]=kw
           del skwords
           #misspelled=sc.unknown([kw])
           #if len(misspelled) == 0:
           #   continue
           #else:
           #   kw=sc.correction(kw)
           #   kwords[i]=kw
       #del skwords

   v=' '.join(kwords)
   keys=[]
   # Go trough a dictonary sorted by len
   #dkeys=list(d.keys())
   #dkeys.sort(key=len, reverse=True)
   for k in d: #dkeys:
       ks=k.split()
       # For single words we want to match
       # them using spaces to make sure that
       # we match the whole word.
       if len(ks)==1:
           kl=[' %s ' % k,'%s ' % k,' %s' % k]
           for kk in kl:
              if kk in v:
                  keys.append(k)
                  classes.append(d[k])
       else:
           if k in v:
               keys.append(k)
               classes.append(d[k])
   classes=set(classes)
   keys=set(keys)
   if len(classes) == 0:
       return 'Unclassified',()
   return ' + '.join(list(classes)),','.join(list(keys))

def is_english(st):
   try:
      st.encode('ascii')
   except UnicodeEncodeError as e:
      print(e)
      return False
   else:
      return True

def classify_search_terms(paths,out_path,dict_path,log_progress=print):
   sc=SpellChecker()
   # Has to match the order of type of files
   # in paths.
   ptypes=['organic','paid','none']
   extra_rows=['Portfolio Classification','Clicks','Impressions','Average position']

   out_wb=Workbook(write_only=True)
   out_ws=out_wb.create_sheet()
 
   cls_dict=load_dict(dict_path)
   
   open_paths=[]
   # First pass: open files,
   # generate simple portfolio classification
   pc={}
   for ip,p in enumerate(paths):
      wb=load_workbook(p, read_only=True)
      ws=wb.active
      open_paths.append(ws)
      for i,row in enumerate(ws.rows,1):
         # Skip first line and headers.
         if i<3:
            continue
         st=row[0].value
         st=' '.join(str(st).lower().split(' '))
         if not is_english(st):
            print('Not english:',repr(st))
            continue
         if st not in pc:
            pc[st]=[]
         # Add this classification only
         # if it's not already there.
         if ptypes[ip] not in pc[st]:
            pc[st].append(ptypes[ip])
   

   none_matches=[0,0,0]
   partial_matches=[0,0,0]
   full_matches=[0,0,0]
   # The number of search terms in each portfolio.
   pt_totals=[0,0,0]
   from_partial_to_full=0
   from_unclass_to_partial=0
   from_unclass_to_full=0
   # Click stats: paid and organic
   clicks=[0,0]
   for ip,ws in enumerate(open_paths):
      i=1
      log_progress('Starting processing rows...')
      for row in ws.rows:
         # Dealing with the first line in input file
         # and the header, but only when processing the first file.
         if ip == 0 and i < 3:
            r=row
            # Add header only at the beginning
            if i == 2:
               r=[row[0].value,'Semantic Classification']+extra_rows
            else:
               # Rewrite the first line from the input
               r=[ _r.value for _r in r ]
            out_ws.append(r)
            i+=1
            continue
         elif i < 3:
            # Skip first lines and headers
            # for the rest of the files.
            i+=1
            continue
         # Starting to process the rest of the data
         # Getting the search term.
         st=row[0].value
         print('Search term:',st)
         if st is None:
            continue
         st=str(st).lower().split(' ')
         sst=' '.join(st)
         if sst not in pc:
            continue
         csst=pc[sst]
         if csst == 'added':
            print('Search term has been already added')
            continue
         # Trying to classify.
         sclsi,tom,um_kws=classify_kws(st, cls_dict)
         # Try spell correct and classify again
         sst=' '.join(st)
         corrected_st=sst
         current_tom=tom
         # Trying to improve those two
         if tom in ('partial','unclassified'):
            print('Trying to improve ', tom)
            # Going trought keywords that didn't
            # have any match
            for ukw in um_kws:
               w=Word(ukw)
               wc=w.correct()
               # If a spell corrected word is 
               # is new let's use it
               if wc != ukw:
                  corrected_st=corrected_st.replace(ukw, wc)
                  continue
               wl=w.lemmatize()
               if wl != ukw:
                  corrected_st=corrected_st.replace(ukw, wl)
                  continue
         # If we've changed the search term let's
         # classify it again
         if corrected_st != sst:
            #print('Corrected st from %s->%s' % (sst, corrected_st))
            corrected_st=corrected_st.split()
            sclsi,tom,um_kws=classify_kws(corrected_st, cls_dict)
            if current_tom == 'unclassified' and tom == 'partial':
               print('Corrected from %s to %s!' % (current_tom, tom))
               st=corrected_st
               from_unclass_to_partial+=1
            elif current_tom in ('unclassified','partial') and tom == 'full':
               print('Corrected from %s to full!' % current_tom)
               st=corrected_st
               if current_tom == 'partial':
                  from_partial_to_full+=1
               if current_tom == 'unclassified':
                  from_unclass_to_full+=1
         if tom == 'full':
            full_matches[ip]+=1
         elif tom == 'partial':
            partial_matches[ip]+=1
         elif tom == 'unclassified':
            none_matches[ip]+=1
         pt_totals[ip]+=1
         pclsi=' | '.join(pc[sst])
         # Match this search term as added
         # to prevent duplicates.  
         pc[sst]='added'
         # Generating the output
         newrow=[row[0].value,sclsi,pclsi]
         if ptypes[ip] == 'organic' and pclsi.find('organic') != -1:
            # Fill in extra rows only from
            # organic file.
            newrow+=[ _r.value for _r in row[1:] ]
            no_clicks=int(row[1].value)
            # Count clicks
            if pclsi.find('paid') != -1:
               clicks[0]+=no_clicks
            if pclsi.find('organic') != -1:
               clicks[1]+=no_clicks
         out_ws.append(newrow)
         if i % 1000 == 0:
            n=datetime.now()
            m='%s:generated %d rows.' % (n,i)
            log_progress(m)
            print(m)
         i+=1
       #print(st, clsi)
   max_rows=i
   print('Total correction from partial to full: %d' % from_partial_to_full)
   print('Total correction from unclass to partial: %d' % from_unclass_to_partial)
   print('Total correction from unclass tofull: %d' % from_unclass_to_full)
   print('Number of clicks', clicks)
   snames=['full','partial','unclassified']
   stats=[full_matches,partial_matches,none_matches]
   out_ws.append(['Total:'])
   # Total stats for each portfolio.
   out_ws.append(['Match type/Portfolio']+['+'.join(ptypes)])
   
   for i,s in enumerate(stats):
      spp='%.2f%%' % ((sum(s)*100)/sum(pt_totals))
      out_ws.append([snames[i]]+[spp])
   out_ws.append([])
   # Per portfolio stats
   out_ws.append(['Per Portfolio:'])
   out_ws.append(['Match type/Portfolio']+ptypes)
   for i,s in enumerate(stats):
      pp=[]
      for i, m in enumerate(s):
         if pt_totals[i] == 0:
            pp.append('0%')
         else:
            pp.append('%.2f%%' % ((m*100)/pt_totals[i])) 
      out_ws.append([snames[i]]+pp)

   out_ws.append(['Classification improvements:'])
   out_ws.append(['Unclassified->Partial',from_unclass_to_partial])
   out_ws.append(['Partial->Full',from_partial_to_full])
   out_ws.append(['Unclassified->Full',from_unclass_to_full])

   # Generate venn diagram
   import tempfile
   # vd=venn2((clicks[0], abs(clicks[0]-clicks[1]),clicks[1]), set_labels = ('Paid', 'Organic'))
   # vdf = tempfile.NamedTemporaryFile()
   # vdf = vdf.name+'.png'
   # plt.savefig(vdf)
   # img = drawing.image.Image(vdf)
   # max_rows=out_ws._max_row+1
   # imgcell='A%d' % max_rows
   # out_ws.add_image(img, imgcell)
   
   log_progress('Saving output file...')
   out_wb.save(out_path) 

## New
from .models import SearchDictionary, SearchPortfolio, SearchTerm, SearchClassification, SearchMatchTotal, SearchMatch
from background_task import background
from django.utils import timezone
import time

def load_dict2():
  st_dict={}
  for item in SearchDictionary.objects.all():
    k=item.name.lower()
    k=' '.join(k.split())
    st_dict[k]=item.attribute
  return st_dict 

@background(schedule=timezone.now())
def classify_search_terms2():  
  # sc=SpellChecker()
  # Has to match the order of type of files
  # in paths.
  ptypes=['organic','paid','none']
  extra_rows=['Portfolio Classification','Clicks','Impressions','Average position']

  out_wb=Workbook(write_only=True)
  out_ws=out_wb.create_sheet()
 
  cls_dict=load_dict2()
   
  open_paths=[]
  # First pass: open files,
  # generate simple portfolio classification
  pc={}  
  for term in SearchTerm.objects.all():
    st = ' '.join(str(term.query).lower().split(' '))
    if not is_english(st):
      print('Not english:',repr(st))
      continue
    if st not in pc:
      pc[st]=[]
      # Add this classification only
      # if it's not already there.
    if term.search_portfolio.name not in pc[st]:
      pc[st].append(term.search_portfolio.name)
  
  organic_portfolio = SearchPortfolio.objects.get(name='organic')
  paid_portfolio = SearchPortfolio.objects.get(name='paid')

  none_matches = {} # {'organic': 0, 'paid': 0, 'none': 3}
  partial_matches = {}
  full_matches = {}
  pt_totals = {} # {'organic': 0, 'paid': 0, 'none': 3}
  for portfolio in SearchPortfolio.objects.all():
    none_matches[portfolio.name] = 0
    partial_matches[portfolio.name] = 0
    full_matches[portfolio.name] = 0
    pt_totals[portfolio.name] = portfolio.searchterm_set.count()

  from_partial_to_full=0
  from_unclass_to_partial=0
  from_unclass_to_full=0
  # Click stats: paid and organic
  clicks=[0,0]

  i=1
  classifications_list = []
  SearchClassification.objects.all().delete()
  for term in SearchTerm.objects.all():
    # Starting to process the rest of the data
    # Getting the search term.
    st = term.query
    print('%s - Search term: %s' % (term.id, st))
    if st is None:
      continue
    st=str(st).lower().split(' ')
    sst=' '.join(st)
    if sst not in pc:
      continue
    csst=pc[sst]
    if csst == 'added':
      print('Search term has been already added')
      continue
    # Trying to classify.
    sclsi,tom,um_kws=classify_kws(st, cls_dict)
    # Try spell correct and classify again
    sst=' '.join(st)
    corrected_st=sst
    current_tom=tom
    # Trying to improve those two
    if tom in ('partial','unclassified'):
      print('Trying to improve ', tom)
      # Going trought keywords that didn't
      # have any match
      for ukw in um_kws:
        w=Word(ukw)
        wc=w.correct()
        # If a spell corrected word is 
        # is new let's use it
        if wc != ukw:
          corrected_st=corrected_st.replace(ukw, wc)
          continue
        wl=w.lemmatize()
        if wl != ukw:
          corrected_st=corrected_st.replace(ukw, wl)
          continue
    # If we've changed the search term let's
    # classify it again
    if corrected_st != sst:
      #print('Corrected st from %s->%s' % (sst, corrected_st))
      corrected_st=corrected_st.split()
      sclsi,tom,um_kws=classify_kws(corrected_st, cls_dict)
      if current_tom == 'unclassified' and tom == 'partial':
        print('Corrected from %s to %s!' % (current_tom, tom))
        st=corrected_st
        from_unclass_to_partial+=1
      elif current_tom in ('unclassified','partial') and tom == 'full':
        print('Corrected from %s to full!' % current_tom)
        st=corrected_st
        if current_tom == 'partial':
          from_partial_to_full+=1
        if current_tom == 'unclassified':
          from_unclass_to_full+=1
    if tom == 'full':
      full_matches[term.search_portfolio.name] += 1
    elif tom == 'partial':
      partial_matches[term.search_portfolio.name] += 1
    elif tom == 'unclassified':
      none_matches[term.search_portfolio.name] += 1

    pclsi=' | '.join(pc[sst])
    # Match this search term as added
    # to prevent duplicates.  
    pc[sst]='added'
    # Generating the output
    newrow=[term.query, sclsi, pclsi]    
    if term.search_portfolio.name == 'organic' and pclsi.find('organic') != -1:
      # Fill in extra rows only from
      # organic file.      
      newrow+=[term.clicks, term.impressions, term.avg_position, term.cost, term.conversions, term.total_conv]
      no_clicks=int(term.clicks)
      # Count clicks
      if pclsi.find('paid') != -1:
         clicks[0]+=no_clicks
      if pclsi.find('organic') != -1:
         clicks[1]+=no_clicks
    out_ws.append(newrow)
    if i % 1000 == 0:
      n=datetime.now()
      m='%s:generated %d rows.' % (n,i)      
      print(m)
    i+=1

    search_classification = SearchClassification(query = term.query, search_dictionary_ids = sclsi, portfolio_ids = pclsi)
    if pclsi.find('organic') != -1:
      search_term = SearchTerm.objects.filter(query=term.query, search_portfolio_id=organic_portfolio.id).first()
      search_classification.organic_clicks = search_term.clicks
      search_classification.organic_impressions = search_term.impressions
      search_classification.organic_avg_position = search_term.avg_position
    if pclsi.find('paid') != -1:
      search_term = SearchTerm.objects.filter(query=term.query, search_portfolio_id=paid_portfolio.id).first()
      search_classification.paid_clicks = search_term.clicks
      search_classification.paid_impressions = search_term.impressions
      search_classification.paid_cost = search_term.cost
      search_classification.paid_conversions = search_term.conversions
      search_classification.paid_total_conv = search_term.total_conv
    print("query - %s" % search_classification.query)
    print("semantic - %s" % search_classification.search_dictionary_ids)
    search_classification.save()    
    # classifications_list.append(search_classification)
    
  # SearchClassification.objects.bulk_create(classifications_list)
  print("SearchClassification records are created.")

  print('*****************')
  print(none_matches)
  print(partial_matches)
  print(full_matches)
  print(pt_totals)

  max_rows=i
  print("max_rows: %s" % max_rows)
  max_rows=SearchTerm.objects.count()
  print("max_rows: %s" % max_rows)
  print('Total correction from partial to full: %d' % from_partial_to_full)
  print('Total correction from unclass to partial: %d' % from_unclass_to_partial)
  print('Total correction from unclass tofull: %d' % from_unclass_to_full)
  print('Number of clicks', clicks)
  
  snames=['full', 'partial', 'unclassified']
  stats=[full_matches, partial_matches, none_matches]
  out_ws.append(['Total:'])
  # Total stats for each portfolio.
  out_ws.append(['Match type/Portfolio']+['+'.join(ptypes)])

  total_match = {}
  match = []
  for i,s in enumerate(stats):
    spp='%.2f%%' % ((sum(s.values())*100)/sum(pt_totals.values()))
    out_ws.append([snames[i]]+[spp])
    total_match[snames[i]] = (sum(s.values())*100)/sum(pt_totals.values())
  out_ws.append([])

  # Per portfolio stats
  out_ws.append(['Per Portfolio:'])
  out_ws.append(['Match type/Portfolio']+ptypes)

  for i,s in enumerate(stats):
    pp = []
    match_t = []
    for item in s.items():
      portfolio = SearchPortfolio.objects.get(name = item[0])
      if pt_totals[item[0]] == 0:
        pp.append('0%')
        match_t.append((portfolio.id, 0))
      else:
        pp.append('%.2f%%' % ((item[1]*100)/pt_totals[item[0]]))
        match_t.append( (portfolio.id, round((item[1]*100)/pt_totals[item[0]], 2)) )
    out_ws.append([snames[i]]+pp)
    match.append(match_t)

  out_ws.append(['Classification improvements:'])
  out_ws.append(['Unclassified->Partial',from_unclass_to_partial])
  out_ws.append(['Partial->Full',from_partial_to_full])
  out_ws.append(['Unclassified->Full',from_unclass_to_full])

  total_match['unclassified_partial'] = from_unclass_to_partial
  total_match['partial_full'] = from_partial_to_full
  total_match['unclassified_full'] = from_unclass_to_full

  SearchMatchTotal.objects.all().delete()
  total = SearchMatchTotal(**total_match)
  total.save()

  match_list = []
  for i in range(0, len(match[0])):
    item = [item[i] for item in match]
    print(item)
    match_list.append(SearchMatch(full = item[0][1], partial = item[1][1], unclassified = item[2][1], search_portfolio_id = item[0][0]))
  SearchMatch.objects.all().delete()
  SearchMatch.objects.bulk_create(match_list)

  # Generate venn diagram 
    
  # out_path = '/home/engineer/Projects/Python/Shilo/classified.xlsx'
  # out_wb.save(out_path)